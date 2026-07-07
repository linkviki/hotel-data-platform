import calendar
import json
import re
import sys
from datetime import datetime
from functools import lru_cache
from pathlib import Path

from openpyxl import load_workbook

from models.status import ImportStatus
from validators.historical_validator import validate_historical_row
from writers.google_sheets import append_historical_monthly, append_import_log


DEFAULT_SAMPLE_PATH = Path("samples/07- July Final P&L - Residence Inn Laval.xlsx")
HOTELS_PATH = Path("config/hotels.json")
OUTPUT_JSON_PATH = Path("output/historical_pl_extracted.json")
REPORT_TYPE = "Historical Monthly P&L"

MONTHS_BY_NUMBER = {index: calendar.month_name[index] for index in range(1, 13)}
MONTH_NUMBER_BY_NAME = {
    calendar.month_name[index].lower(): index for index in range(1, 13)
}
MONTH_NUMBER_BY_ABBREVIATION = {
    calendar.month_abbr[index].lower(): index for index in range(1, 13)
}

MONTH_REPLACEMENTS = [
    (r"\bhiex\b", "holiday inn express"),
]

FIELD_LABELS = {
    "rooms_sold": ["rooms sold"],
    "occupancy_pct": ["occupancy %"],
    "adr": ["adr"],
    "revpar": ["revpar"],
    "room_revenue": ["room revenue", "rooms"],
    "total_revenue": ["total revenue"],
}


def canonicalize_text(value: str) -> str:
    normalized = re.sub(r"[^a-z0-9]+", " ", value.lower())
    return re.sub(r"\s+", " ", normalized).strip()


def normalize_text_for_matching(value: str) -> str:
    normalized = canonicalize_text(value)
    for pattern, replacement in MONTH_REPLACEMENTS:
        normalized = re.sub(pattern, replacement, normalized)
    normalized = re.sub(r"\bfinal\b", " ", normalized)
    normalized = re.sub(r"\bpl\b", " ", normalized)
    normalized = re.sub(r"\bp&l\b", " ", normalized)
    normalized = re.sub(r"\bmonth\b", " ", normalized)
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return normalized


@lru_cache(maxsize=1)
def load_hotel_registry() -> list[dict]:
    payload = json.loads(HOTELS_PATH.read_text(encoding="utf-8-sig"))

    if not isinstance(payload, dict):
        raise ValueError(f"Invalid hotel registry structure in {HOTELS_PATH}")

    registry = []

    for canonical_name, metadata in payload.items():
        if not isinstance(canonical_name, str) or not canonical_name.strip():
            continue

        if not isinstance(metadata, dict):
            continue

        aliases = metadata.get("aliases", [])
        normalized_aliases = []
        if isinstance(aliases, list):
            normalized_aliases = [
                normalize_text_for_matching(alias)
                for alias in aliases
                if isinstance(alias, str) and alias.strip()
            ]

        normalized_name = normalize_text_for_matching(canonical_name)

        registry.append(
            {
                "canonical_name": canonical_name.strip(),
                "normalized_name": normalized_name,
                "tokens": tuple(normalized_name.split()),
                "aliases": normalized_aliases,
            }
        )

    registry.sort(key=lambda item: len(item["tokens"]), reverse=True)

    if not registry:
        raise ValueError(f"No valid hotel names found in {HOTELS_PATH}")

    return registry


def normalize_hotel_name(raw_hotel_name: str | None, workbook_hotel_name: str | None = None) -> str | None:
    candidates = [raw_hotel_name, workbook_hotel_name]

    for candidate in candidates:
        if candidate is None:
            continue

        candidate_text = str(candidate).strip()
        if not candidate_text:
            continue

        normalized_candidate = normalize_text_for_matching(candidate_text)
        candidate_tokens = set(normalized_candidate.split())

        for hotel in load_hotel_registry():
            canonical_name = hotel["canonical_name"]
            canonical_normalized = hotel["normalized_name"]
            canonical_tokens = set(hotel["tokens"])

            if normalized_candidate == canonical_normalized:
                return canonical_name

            if normalized_candidate in canonical_normalized or canonical_normalized in normalized_candidate:
                return canonical_name

            if hotel["aliases"] and normalized_candidate in hotel["aliases"]:
                return canonical_name

            if canonical_tokens.issubset(candidate_tokens):
                return canonical_name

    return raw_hotel_name or workbook_hotel_name


def parse_month_from_text(value: str) -> int | None:
    normalized = canonicalize_text(value)

    month_match = re.search(
        r"\b(" + "|".join(calendar.month_name[1:]) + r"|" + "|".join(calendar.month_abbr[1:]) + r")\b",
        value,
        flags=re.IGNORECASE,
    )
    if month_match:
        month_text = month_match.group(1).lower()
        return MONTH_NUMBER_BY_NAME.get(month_text) or MONTH_NUMBER_BY_ABBREVIATION.get(month_text)

    number_match = re.match(r"^(0?[1-9]|1[0-2])\b", normalized)
    if number_match:
        return int(number_match.group(1))

    return None


def parse_year_from_text(value: str) -> int | None:
    match = re.search(r"\b(20\d{2})\b", value)
    if match:
        return int(match.group(1))
    return None


def extract_hotel_candidate_from_filename(file_path: Path) -> str | None:
    stem = file_path.stem
    stem = re.sub(r"^\s*\d{1,2}[.\-]?\s*", "", stem)
    stem = re.sub(r"\b(?:January|February|March|April|May|June|July|August|September|October|November|December)\b", "", stem, flags=re.IGNORECASE)
    stem = re.sub(r"\b\d{4}\b", "", stem)
    stem = re.sub(r"\b(?:Final\s+)?P&L\b", "", stem, flags=re.IGNORECASE)
    stem = re.sub(r"\bP&L\b", "", stem, flags=re.IGNORECASE)
    stem = re.sub(r"\s+", " ", stem).strip(" -_")
    return stem or None


def detect_month_year(workbook, file_path: Path) -> tuple[int | None, int | None]:
    month_number = parse_month_from_text(file_path.stem)
    year = parse_year_from_text(file_path.stem)

    sheet = workbook[workbook.sheetnames[0]]

    if month_number is None or year is None:
        for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 20)):
            for cell in row:
                if isinstance(cell.value, str):
                    if month_number is None:
                        month_number = parse_month_from_text(cell.value) or month_number
                    if year is None:
                        year = parse_year_from_text(cell.value) or year

            if month_number is not None and year is not None:
                break

    return month_number, year


def load_workbook_sheet(excel_path: Path):
    workbook = load_workbook(excel_path, data_only=True)
    return workbook, workbook[workbook.sheetnames[0]]


def find_row_value(sheet, target_labels: list[str]):
    normalized_targets = [normalize_text_for_matching(label) for label in target_labels]

    for row in sheet.iter_rows():
        row_labels = []
        for cell in row:
            if isinstance(cell.value, str):
                row_labels.append(normalize_text_for_matching(cell.value))

        if not row_labels:
            continue

        match_found = any(target == label for target in normalized_targets for label in row_labels)
        if not match_found:
            continue

        for cell in row:
            value = cell.value
            if isinstance(value, (int, float)):
                return value

    return None


def normalize_numeric_value(field_name: str, value):
    if value is None:
        return None

    if field_name == "occupancy_pct" and isinstance(value, (int, float)) and 0 <= value <= 1:
        return round(value * 100, 2)

    if isinstance(value, float) and value.is_integer():
        return int(value)

    return value


def extract_historical_pl(excel_path: Path) -> dict:
    workbook, sheet = load_workbook_sheet(excel_path)
    month_number, year = detect_month_year(workbook, excel_path)

    file_hotel_candidate = extract_hotel_candidate_from_filename(excel_path)
    workbook_hotel_candidate = None

    for cell in sheet["A"][:5]:
        if isinstance(cell.value, str) and cell.value.strip():
            workbook_hotel_candidate = cell.value
            break

    hotel_name = normalize_hotel_name(file_hotel_candidate, workbook_hotel_candidate)

    row = {
        "hotel_name": hotel_name,
        "year": year,
        "month_number": month_number,
        "month_name": MONTHS_BY_NUMBER.get(month_number),
        "rooms_sold": normalize_numeric_value("rooms_sold", find_row_value(sheet, FIELD_LABELS["rooms_sold"])),
        "occupancy_pct": normalize_numeric_value("occupancy_pct", find_row_value(sheet, FIELD_LABELS["occupancy_pct"])),
        "adr": normalize_numeric_value("adr", find_row_value(sheet, FIELD_LABELS["adr"])),
        "revpar": normalize_numeric_value("revpar", find_row_value(sheet, FIELD_LABELS["revpar"])),
        "room_revenue": normalize_numeric_value("room_revenue", find_row_value(sheet, FIELD_LABELS["room_revenue"])),
        "total_revenue": normalize_numeric_value("total_revenue", find_row_value(sheet, FIELD_LABELS["total_revenue"])),
        "source_file_name": excel_path.name,
        "import_time": datetime.now().isoformat(timespec="seconds"),
        "status": ImportStatus.RAW_EXTRACTED,
        "notes": "",
    }

    return row


def safe_append_import_log(data: dict, action: str) -> bool:
    try:
        append_import_log(data, action=action)
        return True
    except Exception as exc:
        print(f"Failed to write Import_Log entry: {exc}")
        return False


def safe_append_historical_monthly(rows: list[dict]) -> dict | None:
    try:
        return append_historical_monthly(rows)
    except Exception as exc:
        print(f"Failed to write Historical_Monthly rows: {exc}")
        return None


def resolve_input_files(argv: list[str]) -> list[Path]:
    if not argv:
        return [DEFAULT_SAMPLE_PATH]

    resolved_files = []

    for argument in argv:
        candidate = Path(argument)
        if candidate.exists():
            resolved_files.append(candidate)
            continue

        sample_matches = list(Path("samples").rglob(candidate.name))
        if sample_matches:
            resolved_files.append(sample_matches[0])
            continue

        resolved_files.append(candidate)

    missing_files = [path for path in resolved_files if not path.exists()]
    if missing_files:
        for path in missing_files:
            print(f"File not found: {path}")
        raise SystemExit(1)

    return resolved_files


def process_historical_file(excel_path: Path):
    try:
        raw_result = extract_historical_pl(excel_path)
    except Exception as exc:
        failure_data = {
            "import_time": datetime.now().isoformat(timespec="seconds"),
            "hotel_name": None,
            "report_type": REPORT_TYPE,
            "source_file_name": excel_path.name,
            "status": ImportStatus.EXTRACTION_FAILED,
            "notes": str(exc),
        }

        safe_append_import_log(failure_data, action=ImportStatus.EXTRACTION_FAILED)
        print("Historical extraction failed.")
        print(f"- {exc}")
        return

    print(json.dumps(raw_result, indent=4))

    validation_errors = validate_historical_row(raw_result)

    if validation_errors:
        safe_append_import_log(
            {
                "import_time": raw_result.get("import_time"),
                "hotel_name": raw_result.get("hotel_name"),
                "report_type": REPORT_TYPE,
                "source_file_name": raw_result.get("source_file_name"),
                "status": ImportStatus.VALIDATION_FAILED,
                "notes": "; ".join(validation_errors[:10]),
            },
            action=ImportStatus.VALIDATION_FAILED,
        )

        print("Historical validation failed.")
        for error in validation_errors[:10]:
            print(f"- {error}")
        return

    validated_result = dict(raw_result)
    validated_result["status"] = ImportStatus.VALIDATED
    validated_result["notes"] = ""

    print(json.dumps(validated_result, indent=4))

    result = safe_append_historical_monthly([validated_result])

    if result is None:
        print("Historical Monthly append failed.")
        return

    action = ImportStatus.IMPORTED if result["appended"] else ImportStatus.DUPLICATE_SKIPPED
    status = ImportStatus.VALIDATED if result["appended"] else ImportStatus.DUPLICATE_SKIPPED

    safe_append_import_log(
        {
            "import_time": raw_result.get("import_time"),
            "hotel_name": raw_result.get("hotel_name"),
            "report_type": REPORT_TYPE,
            "source_file_name": raw_result.get("source_file_name"),
            "status": status,
            "notes": f"Appended: {result['appended']}, Skipped: {result['skipped']}",
        },
        action=action,
    )

    print(f"Historical Monthly appended: {result['appended']}")
    print(f"Historical Monthly skipped: {result['skipped']}")


if __name__ == "__main__":
    for file_path in resolve_input_files(sys.argv[1:]):
        process_historical_file(file_path)
